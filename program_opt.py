import openpyxl
import sys
import os
from docx import Document
from docx.shared import RGBColor

document = Document()

def load_files():
    argc = len(sys.argv)
    if argc > 1:
        assortment_price = sys.argv[1]
    if argc > 2:
        price_list = sys.argv[2]
    if argc > 3:
        catalog_ostatkov = sys.argv[3]
    if argc > 4:
        shablon_ostatkov = sys.argv[4]
    
    if not os.path.isfile(catalog_ostatkov):
        print("Файл не найден.")
        return 0;
    if not os.path.isfile(shablon_ostatkov):
        print("Файл не найден.")
        return 0;
    if not os.path.isfile(price_list):
        print("Файл не найден.")
        return 0;
    if not os.path.isfile(assortment_price):
        print("Файл не найден.")
        return 0;
    
    print("Загрузка файлов...")
    assortment_price_book = openpyxl.load_workbook(assortment_price)
    assortment_price_table = assortment_price_book.worksheets[1]
    price_list_table = openpyxl.load_workbook(price_list)
    price_list_table = price_list_table.active
    catalog_ostatkov_table = openpyxl.load_workbook(catalog_ostatkov)
    catalog_ostatkov_table = catalog_ostatkov_table.active
    shablon_ostatkov_book = openpyxl.load_workbook(shablon_ostatkov)
    shablon_ostatkov_table = shablon_ostatkov_book.active
    first_load(assortment_price_table, assortment_price_book, assortment_price, price_list_table, catalog_ostatkov_table, shablon_ostatkov_table, shablon_ostatkov_book, shablon_ostatkov)



def calculate_price(articul, price_value):
    if price_value is None:
        print("\nЗначение \"Цена\" у наименования {} пустое, пожалуйста заполните вручную. Установлено стандартное значение 0.".format(articul))
        return 0
    elif price_value < 1000:
        price_value += (60 * price_value) / 100
    elif price_value >= 1000 and price_value < 10000:
        price_value += (57 * price_value) / 100
    else:
        price_value += (50 * price_value) / 100
    return round(price_value)

def first_load(assortment, assortment_book, assortment_file, price_list, catalog_ostatkov, shablon_ostatkov, shablon_ostatkov_book, shablon_ostatkov_file):
    print("Обновление данных...")
    max_row = assortment.max_row
    max_row2 = price_list.max_row
    max_row3 = catalog_ostatkov.max_row
    max_row4 = shablon_ostatkov.max_row
    articul1 = None
    articul2 = None
    articul3 = None
    remainder_value = None
    for row_assortment in range(4, max_row + 1):
        row_num = row_assortment
        articul1 = assortment.cell(row=row_num, column=3).value
        for row_price_list in range(2, max_row2 + 1):
            articul2 = price_list.cell(row=row_price_list, column=4).value
            if articul1 == articul2:
                price_value = calculate_price(articul2, price_list.cell(row=row_price_list, column=13).value)
                if price_list.cell(row=row_price_list, column=11).value == 'Да':
                    for row_catalog_ostatkov in range(2, max_row3 + 1):
                        articul3 = catalog_ostatkov.cell(row=row_catalog_ostatkov, column=4).value
                        if articul2 == articul3:
                            remainder_value = catalog_ostatkov.cell(row=row_catalog_ostatkov, column=7).value
                            if remainder_value is None:
                                remainder_value = 0
                            break
                else:
                    remainder_value = 0
                for row_shablon_ostatkov in range(4, max_row4 + 1):
                    if shablon_ostatkov.cell(row=row_shablon_ostatkov, column=3).value == articul3:
                        shablon_ostatkov.cell(row=row_shablon_ostatkov, column=5).value = remainder_value
                        break
                assortment.cell(row=row_num, column=5).value = price_value
                assortment.cell(row=row_num, column=6).value = ((price_value * 0.1) + price_value)
                assortment.cell(row=row_num, column=10).value = 1
                assortment.cell(row=row_num, column=11).value = "%"
                assortment.cell(row=row_num, column=7).value = 6
                break
        progress = (row_num / max_row) * 100
        print(str(round(progress)) + "%", end="\r")
    assortment_book.save(assortment_file)
    shablon_ostatkov_book.save(shablon_ostatkov_file)
    print("Обновление данных завершено.")


load_files()
